# -*- coding: utf-8 -*-

# =================================================================================
# osintgpt
#
# Author: @estebanpdl
#
# File: tabular.py
# Description: Rows and records into documents. These formats carry many fields
#   and only some are content, so the mapping is required rather than guessed.
# =================================================================================

# import modules
import csv
import json

# import submodules
from pathlib import Path

# type hints
from typing import Any, Dict, Iterator, List, Optional, Union

from .documents import Document, FieldMapping, document_from_record, value_at

# Rows sampled when describing a file's fields. Enough to tell a body of text
# from an identifier without reading a corpus to answer a question about it.
SAMPLE_ROWS = 50


# UnmappedSourceError class
class UnmappedSourceError(ValueError):
    '''
    Raised when a structured file is loaded without being told which fields
    carry its content.
    '''


# describe what a structured file contains
def describe_fields(path: Union[str, Path]) -> Dict[str, dict]:
    '''
    Sample a file and report what each field looks like.

    Feeds the dry run, where an operator turns a description into a mapping.
    Reports what the values are; it never decides what they are for.

    Args:
        path (Union[str, Path]): File to sample.

    Returns:
        Dict[str, dict]: Per field, its average text length, how many of the \
            sampled values were filled, whether they were unique, and one \
            example.
    '''
    records = list(_sample(Path(path)))
    fields: Dict[str, dict] = {}

    for name in _field_names(records):
        values = [value_at(record, name) for record in records]
        filled = [str(v).strip() for v in values if v is not None and str(v).strip()]
        fields[name] = {
            'filled': len(filled),
            'sampled': len(values),
            'average_length': round(
                sum(len(v) for v in filled) / len(filled)
            ) if filled else 0,
            'unique': len(set(filled)) == len(filled) and len(filled) > 1,
            'example': filled[0][:120] if filled else ''
        }

    return fields


# load a structured file as documents
def load_records(
    path: Union[str, Path], mapping: FieldMapping
) -> Iterator[Document]:
    '''
    Read a tabular or nested file into one document per record.

    Args:
        path (Union[str, Path]): File to read.
        mapping (FieldMapping): Which fields play which role.

    Raises:
        UnmappedSourceError: If no content field was named. Indexing every \
            field would bury the content under identifiers that repeat across \
            every record, and the result looks populated rather than broken.

    Yields:
        Document: One per record that had content.
    '''
    path = Path(path)
    if not mapping.is_set:
        raise UnmappedSourceError(
            f'{path.name} is structured data; name the field or fields '
            'carrying its content before indexing it. Available fields: '
            + ', '.join(describe_fields(path)) or '(none found)'
        )

    source_ref = path.as_posix()
    for position, record in enumerate(_records(path, mapping)):
        document = document_from_record(record, mapping, source_ref, position)
        if document is not None:
            yield document


# records from any supported structured format
def _records(path: Path, mapping: FieldMapping) -> Iterator[Any]:
    suffix = path.suffix.lower()
    if suffix == '.csv':
        return _csv_records(path)
    if suffix in ('.xlsx', '.xlsm'):
        return _excel_records(path)
    if suffix in ('.json', '.jsonl', '.ndjson'):
        return _json_records(path, mapping)

    raise ValueError(f'{path.name}: not a structured format osintgpt reads')


def _csv_records(path: Path) -> Iterator[dict]:
    # newline='' is required by csv; utf-8-sig drops the BOM that spreadsheet
    # exports carry, which would otherwise corrupt the first column's name.
    with open(path, newline='', encoding='utf-8-sig', errors='replace') as handle:
        yield from csv.DictReader(handle)


def _excel_records(path: Path) -> Iterator[dict]:
    from openpyxl import load_workbook

    # read_only streams rather than loading the sheet; data_only takes a
    # formula's last computed value instead of its source.
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return

        names = [str(name) if name is not None else '' for name in header]
        for row in rows:
            yield {
                name: value
                for name, value in zip(names, row)
                if name
            }
    finally:
        workbook.close()


def _json_records(path: Path, mapping: FieldMapping) -> Iterator[Any]:
    text = path.read_text(encoding='utf-8', errors='replace')

    if path.suffix.lower() in ('.jsonl', '.ndjson'):
        for line in text.splitlines():
            line = line.strip()
            if line:
                yield json.loads(line)
        return

    parsed = json.loads(text)
    if mapping.records:
        parsed = value_at(parsed, mapping.records)

    if isinstance(parsed, list):
        yield from parsed
    elif isinstance(parsed, dict):
        yield parsed


# a bounded sample, for describing rather than loading
def _sample(path: Path) -> Iterator[Any]:
    empty = FieldMapping()
    try:
        for index, record in enumerate(_records(path, empty)):
            if index >= SAMPLE_ROWS:
                return
            yield record
    except (ValueError, KeyError):
        return


def _field_names(records: List[Any]) -> List[str]:
    '''
    Field names in first-seen order, descending into nested records one level
    at a time so a dotted path is offered rather than an opaque object.
    '''
    names: List[str] = []

    def walk(record: Any, prefix: str = '') -> None:
        if not isinstance(record, dict):
            return
        for key, value in record.items():
            path = f'{prefix}{key}'
            if isinstance(value, dict):
                walk(value, f'{path}.')
            elif path not in names:
                names.append(path)

    for record in records:
        walk(record)

    return names
